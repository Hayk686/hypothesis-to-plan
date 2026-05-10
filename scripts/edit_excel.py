#!/usr/bin/env python3
"""Excel editor for .xlsx and .xlsm: list sheets, read, write cell, export CSV.

Emits a single JSON line on stdout (status/files/count, plus optional
content/sheets/error fields). .xls (legacy) is not supported by openpyxl.
"""

import argparse
import csv
import io
import json
import os
import sys

try:
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")
except AttributeError:
    pass

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter
    from openpyxl.utils.cell import coordinate_from_string, column_index_from_string
except ImportError:
    print(json.dumps({
        "status": "error",
        "error": "openpyxl is not installed. Run: pip install openpyxl",
        "files": [],
        "count": 0,
    }, ensure_ascii=False))
    sys.exit(2)

OUTPUT_DIR = os.path.join("output", "documents")
SUPPORTED_EXT = {".xlsx", ".xlsm"}


def emit(payload: dict, code: int = 0) -> None:
    print(json.dumps(payload, ensure_ascii=False, default=str))
    sys.exit(code)


def strip_file_wrapper(path: str) -> str:
    p = path.strip()
    if p.startswith("[file:") and p.endswith("]"):
        p = p[6:-1]
    return p


def check_input(path: str) -> str:
    src = strip_file_wrapper(path)
    if not os.path.isfile(src):
        emit({"status": "error", "error": f"file not found: {src}", "files": [], "count": 0}, 1)
    ext = os.path.splitext(src)[1].lower()
    if ext not in SUPPORTED_EXT:
        emit({
            "status": "error",
            "error": f"unsupported extension: {ext}. Supported: .xlsx, .xlsm. Convert .xls to .xlsx first.",
            "files": [], "count": 0,
        }, 1)
    return src


def pick_sheet(wb, sheet_name: str | None):
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            emit({
                "status": "error",
                "error": f"sheet not found: {sheet_name}. Available: {wb.sheetnames}",
                "files": [], "count": 0,
            }, 1)
        return wb[sheet_name]
    return wb.active


def sheet_to_csv(ws, max_rows: int | None) -> str:
    buf = io.StringIO()
    writer = csv.writer(buf, lineterminator="\n")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if max_rows is not None and i >= max_rows:
            break
        writer.writerow(["" if v is None else v for v in row])
    return buf.getvalue()


def cmd_list(args) -> None:
    src = check_input(args.input)
    wb = load_workbook(src, data_only=True, read_only=True, keep_links=False)
    emit({"status": "ok", "sheets": wb.sheetnames, "files": [], "count": 0})


def cmd_read(args) -> None:
    src = check_input(args.input)
    wb = load_workbook(src, data_only=True, read_only=True, keep_links=False)
    ws = pick_sheet(wb, args.sheet)
    content = sheet_to_csv(ws, args.max_rows)
    emit({"status": "ok", "sheet": ws.title, "content": content, "files": [], "count": 0})


def cmd_write(args) -> None:
    src = check_input(args.input)
    keep_vba = src.lower().endswith(".xlsm")
    wb = load_workbook(src, keep_vba=keep_vba)
    ws = pick_sheet(wb, args.sheet)
    try:
        coordinate_from_string(args.cell)
    except Exception:
        emit({"status": "error", "error": f"invalid cell coordinate: {args.cell}", "files": [], "count": 0}, 1)
    value = args.value
    if args.value_type == "number":
        try:
            value = float(args.value) if "." in args.value else int(args.value)
        except ValueError:
            emit({"status": "error", "error": f"value is not a number: {args.value}", "files": [], "count": 0}, 1)
    elif args.value_type == "bool":
        value = args.value.strip().lower() in ("1", "true", "yes")
    ws[args.cell] = value

    if args.output:
        out = args.output
        os.makedirs(os.path.dirname(os.path.abspath(out)) or ".", exist_ok=True)
    else:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        base = os.path.basename(src)
        out = os.path.join(OUTPUT_DIR, base)
    wb.save(out)
    emit({"status": "ok", "files": [os.path.abspath(out)], "count": 1, "sheet": ws.title, "cell": args.cell})


def cmd_to_csv(args) -> None:
    src = check_input(args.input)
    wb = load_workbook(src, data_only=True, read_only=True, keep_links=False)
    ws = pick_sheet(wb, args.sheet)
    if args.output:
        out = args.output
        os.makedirs(os.path.dirname(os.path.abspath(out)) or ".", exist_ok=True)
    else:
        os.makedirs(OUTPUT_DIR, exist_ok=True)
        base = os.path.splitext(os.path.basename(src))[0]
        out = os.path.join(OUTPUT_DIR, f"{base}__{ws.title}.csv")
    with open(out, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        for row in ws.iter_rows(values_only=True):
            writer.writerow(["" if v is None else v for v in row])
    emit({"status": "ok", "files": [os.path.abspath(out)], "count": 1, "sheet": ws.title})


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Excel (.xlsx/.xlsm) reader/editor")
    sub = parser.add_subparsers(dest="action", required=True)

    p_list = sub.add_parser("list", help="List sheet names")
    p_list.add_argument("--input", required=True)
    p_list.set_defaults(func=cmd_list)

    p_read = sub.add_parser("read", help="Read a sheet as CSV text")
    p_read.add_argument("--input", required=True)
    p_read.add_argument("--sheet", help="Sheet name. Defaults to the active sheet.")
    p_read.add_argument("--max-rows", type=int, dest="max_rows")
    p_read.set_defaults(func=cmd_read)

    p_write = sub.add_parser("write", help="Set a cell value")
    p_write.add_argument("--input", required=True)
    p_write.add_argument("--sheet", help="Sheet name. Defaults to the active sheet.")
    p_write.add_argument("--cell", required=True, help="A1-style coordinate, e.g. B7")
    p_write.add_argument("--value", required=True)
    p_write.add_argument("--value-type", dest="value_type", choices=["string", "number", "bool"], default="string")
    p_write.add_argument("--output", help="Optional output path. Defaults to output\\documents\\<name>")
    p_write.set_defaults(func=cmd_write)

    p_csv = sub.add_parser("to-csv", help="Export a sheet as a .csv file")
    p_csv.add_argument("--input", required=True)
    p_csv.add_argument("--sheet", help="Sheet name. Defaults to the active sheet.")
    p_csv.add_argument("--output", help="Optional output path. Defaults to output\\documents\\<name>__<sheet>.csv")
    p_csv.set_defaults(func=cmd_to_csv)

    return parser


def main() -> None:
    args = build_parser().parse_args()
    try:
        args.func(args)
    except Exception as e:
        emit({"status": "error", "error": str(e), "files": [], "count": 0}, 1)


if __name__ == "__main__":
    main()
